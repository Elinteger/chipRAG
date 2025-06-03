"""
Pipeline that creates and saves a pre-formatted Excel file comparing the Maximum Residue Limit (MRL) values for specified pesticides and foods.
"""
import logging
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from .postgres_utils import query_database
from .chiprag_modules import extract_relevant_values, get_fitting_pesticides, compare_values
from .postgres_utils import get_pesticide_data


def create_comparison(
        keywords: list[str],
        output_path: str
) -> pd.DataFrame:
    """
    Generates a formatted Excel sheet comparing Chinese and European Maximum Residue Limit (MRL) values for specified pesticides and foods.

    Args:
        keywords (list[str] | str): Required to know which pesticides/foods should be compared.
        output_path (str): Path where the excel output should be saved to. Defaults to "output.xlsx" in the working directory.

    Returns:
        pd.DataFrame: DataFrame with the exact same output as is in the excel sheet.
    """
    logging.info("-- Creating comparison --")
    # get values which are relevant for comparison
    chi_values = _get_chi_values(keywords)
    if chi_values.empty:
        logging.info("No values found, aborting comparison.")
        return chi_values
    logging.info("Got chinese values.")
    eu_values, bridge_dict = _get_eu_values(chi_values)
    logging.info("Got european values.")
    # create comparison
    comparison = compare_values(chi_values, eu_values, bridge_dict)
    logging.info("Created comparison.")
    # save as formatted excel
    formatted_comparsion = _render_to_xlsx(comparison, output_path)
    logging.info(f"Saved formatted excel sheet at {output_path}.")

    return formatted_comparsion


def _get_chi_values(
        keywords: list[str]
) -> pd.DataFrame:
    """
    Helper function that retrieves all information for a given list of keywords from the GBs containing Chinese pesticide Maximum Residue Limit data.

    Args:
        keywords (list[str]): List of pesticides and foods to gather information for. Keywords must exactly match the English translations of the GB.

    Returns:
        pd.DataFrame: DataFrame containing all relevant information with columns: 'pesticide', 'food', and 'mrl'.
    """
    # get all entries from the database
    list = query_database(keywords)
    if len(list) == 0:
         logging.warning("Couldn't find any values in the database fitting the users request. Check request and database accordingly.")
         return pd.DataFrame(list)
    
    # filter out and return the relevant ones
    return extract_relevant_values(keywords, list)


def _get_eu_values(
        chi_values: pd.DataFrame
) -> tuple[pd.DataFrame, dict]:
    """
    Helper function that retrieves all relevant information based on Chinese pesticide values.

    Args:
        chi_values (pd.DataFrame): DataFrame containing all relevant Chinese values.

    Returns:
        tuple[pd.DataFrame, dict]: A tuple with a DataFrame of European values corresponding to the Chinese entries, 
        and a bridge dictionary mapping Chinese pesticide names to European pesticide names.
    """
    # get fitting eu_pesticides, dict that acts as a bridge between chi_values and eu_values
    eu_pesticide_dict = get_fitting_pesticides(chi_values)
    # get all data regarding those pesticides
    eu_pesticides_list = [item for sublist in eu_pesticide_dict.values() for item in sublist]
    eu_values = get_pesticide_data(eu_pesticides_list)
    # build dataframe
    rows = []
    for chi_pest, fitting_eu_pest in eu_pesticide_dict.items():
        if len(fitting_eu_pest) == 0:
            rows.append({
                    'chi_pesticide': chi_pest,
                    'eu_pesticide': "/",
                    'food': "/",
                    'mrl': "/"
                    })
            continue
        
        for eu_pest_key in fitting_eu_pest:
            if eu_pest_key in eu_values:
                for items in eu_values[eu_pest_key]:
                    rows.append({
                    'chi_pesticide': chi_pest,
                    'eu_pesticide': items[0],
                    'food': items[1],
                    'mrl': items[2]
                    })

    eu_df = pd.DataFrame(rows)

    return eu_df, eu_pesticide_dict


def _render_to_xlsx(
        comparsion_df: pd.DataFrame, 
        output_path: str
) -> pd.DataFrame:
    """
    Helper function that formats a given DataFrame, exports it as a formatted Excel sheet, and saves it to the specified path.

    Args:
        comparison_df (pd.DataFrame): DataFrame containing the comparison data.
        output_path (str): File path where the Excel file will be saved.

    Returns:
        pd.DataFrame: The input DataFrame with the 'note' and 'valid_mrl' columns swapped for improved readability.
    """
    # change order
    comparsion_df = comparsion_df[[
    'chi_pesticide', 'eu_pesticide',
    'chi_food', 'eu_food',
    'chi_mrl', 'eu_mrl',
    'valid_mrl', 'note'
    ]]

    # turn DataFrame to excel and format
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        comparsion_df.to_excel(writer, index=False, sheet_name="GB Comparison")
        #FIXME: is workbook necessary?
        # workbook = writer.book
        worksheet = writer.sheets["GB Comparison"]

        # set column widths based on max content length
        for col_idx, col in enumerate(comparsion_df.columns, start=1):
            max_length = max(
                comparsion_df[col].astype(str).map(len).max(),
                len(col)
            )
            worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        # make selected columns bold (headers and cells)
        bold_cols = {'eu_pesticide', 'chi_food', 'valid_mrl'}
        bold_font = Font(bold=True)

        for col_idx, col in enumerate(comparsion_df.columns, start=1):
            if col in bold_cols:
                # header bold
                worksheet.cell(row=1, column=col_idx).font = bold_font
                # cell values bold
                for row_idx in range(2, len(comparsion_df) + 2):
                    worksheet.cell(row=row_idx, column=col_idx).font = bold_font

        # conditional formatting for "note" column
        #FIXME: add red for rows where no eu pesticide could be found
        note_col_idx = comparsion_df.columns.get_loc('note') + 1
        for row_idx, note in enumerate(comparsion_df['note'], start=2):
            cell = worksheet.cell(row=row_idx, column=note_col_idx)
            if pd.isna(note) or note == "No Note.":
                # no change
                continue  
            elif "Category." in str(note):
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # light green
            else:
                cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # light yellow

    return comparsion_df


if __name__ == "__main__":
    create_comparison()
